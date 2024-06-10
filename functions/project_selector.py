import tkinter as tk
from tkinter import ttk
import os
import datetime
from openpyxl import load_workbook

class ProjectSelector(tk.Toplevel):
    def __init__(self, app, title, excel_manager):
        super().__init__(app.master)
        self.app = app
        self.title(title)
        self.geometry('600x400')
        self.configure(bg="#333333")
        self.excel_manager = excel_manager

        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.update_list)

        style = ttk.Style()
        style.configure('TLabel', background="#333333", foreground="white", font=('Helvetica', 10))
        style.configure('TEntry', font=('Helvetica', 10))
        style.configure('TTreeview', font=('Helvetica', 10), background="#333333", foreground="white", fieldbackground="#333333")
        style.configure('TTreeview.Heading', font=('Helvetica', 10, 'bold'), background="#333333", foreground="white")

        ttk.Label(self, text="Search:", style='TLabel').pack(pady=5)
        self.search_entry = ttk.Entry(self, textvariable=self.search_var, width=50, style='TEntry')
        self.search_entry.pack(pady=5)

        self.tree = ttk.Treeview(self, columns=("ID", "Project Name", "Category"), show="headings", height=10)
        self.tree.heading("ID", text="ID")
        self.tree.heading("Project Name", text="Project Name")
        self.tree.heading("Category", text="Category")
        self.tree.pack(expand=True, fill="both", pady=10)

        self.tree.bind("<Double-1>", self.on_double_click)

        self.load_projects()
        self.update_list()

    def load_projects(self):
        self.projects = set()
        current_year = datetime.datetime.now().year
        for month in range(1, 13):
            month_file = os.path.join(self.excel_manager.base_directory, str(current_year), f"{month}.xlsx")
            if os.path.exists(month_file):
                wb = load_workbook(month_file)
                for sheet in wb.sheetnames:
                    if sheet != "Total":
                        ws = wb[sheet]
                        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                            if row[1] and row[2]:
                                self.projects.add((row[1], row[2]))

        self.projects = [{"id": i + 1, "name": project[0], "category": project[1]} for i, project in enumerate(self.projects)]

    def update_list(self, *args):
        search_term = self.search_var.get().lower()
        filtered_projects = [project for project in self.projects if search_term in project["name"].lower()]

        for i in self.tree.get_children():
            self.tree.delete(i)

        for project in filtered_projects[:10]:  # Display maximum 10 projects
            self.tree.insert("", "end", values=(project["id"], project["name"], project["category"]))

    def on_double_click(self, event):
        selected_item = self.tree.selection()[0]
        project_name = self.tree.item(selected_item, "values")[1]
        project_category = self.tree.item(selected_item, "values")[2]
        self.app.activity_name.delete(0, tk.END)
        self.app.activity_name.insert(0, project_name)
        self.app.category.set(project_category)
        self.destroy()
