import os
import datetime
from openpyxl import Workbook, load_workbook

class ExcelManager:
    def __init__(self):
        self.base_directory = "Activitate"
        self.ensure_directory_structure()

    def ensure_directory_structure(self):
        current_year = datetime.datetime.now().year
        year_folder = os.path.join(self.base_directory, str(current_year))

        if not os.path.exists(self.base_directory):
            os.makedirs(self.base_directory)
            print(f"Created directory: {self.base_directory}")

        if not os.path.exists(year_folder):
            os.makedirs(year_folder)
            print(f"Created directory: {year_folder}")

        for month in range(1, 13):
            month_file = os.path.join(year_folder, f"{month}.xlsx")
            if not os.path.exists(month_file):
                self.create_month_file(month_file, month)

    def create_month_file(self, filepath, month):
        wb = Workbook()
        # Create a sheet for totals
        total_sheet = wb.active
        total_sheet.title = "Total"
        total_sheet.append(["Date", "Total Time Worked"])

        # Get the number of days in the month
        year = int(filepath.split(os.sep)[-2])
        num_days = self.days_in_month(year, month)

        # Create a row for each day in the "Total" sheet
        for day in range(1, num_days + 1):
            date = datetime.date(year, month, day).strftime("%d/%m/%Y")
            total_sheet.append([date, "00:00:00"])

        # Create a sheet for each day
        for day in range(1, num_days + 1):
            ws = wb.create_sheet(title=f"Day {day}")
            headers = ["ID", "Activity Name", "Category", "Start Time", "End Time", "Duration", "Percentage"]
            ws.append(headers)

        # Save the workbook
        wb.save(filepath)
        print(f"Created monthly log file with daily sheets: {filepath}")

    def days_in_month(self, year, month):
        # Return the number of days in the given month and year
        return (datetime.date(year, month % 12 + 1, 1) - datetime.date(year, month, 1)).days

    def log_activity(self, activity_name, category, start_time, end_time, duration):
        current_date = datetime.datetime.now()
        year = current_date.year
        month = current_date.month
        day = current_date.day
        month_file = os.path.join(self.base_directory, str(year), f"{month}.xlsx")

        if not os.path.exists(month_file):
            self.create_month_file(month_file, month)

        wb = load_workbook(month_file)
        sheet_name = f"Day {day}"
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            headers = ["ID", "Activity Name", "Category", "Start Time", "End Time", "Duration", "Percentage"]
            ws.append(headers)
        else:
            ws = wb[sheet_name]

        new_id = self.get_next_id(ws)
        duration_str = self._format_time(duration)
        ws.append([new_id, activity_name, category, start_time, end_time, duration_str, ""])

        self.update_total_sheet(wb, current_date, duration)
        self.update_percentages(wb, sheet_name)

        wb.save(month_file)

    def get_next_id(self, ws):
        ids = [cell.value for cell in ws['A'] if isinstance(cell.value, int)]
        return max(ids) + 1 if ids else 1

    def calculate_percentage(self, ws, duration_str):
        total_duration = datetime.timedelta()
        for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
            duration = row[5]
            if duration:
                total_duration += datetime.datetime.strptime(duration, "%H:%M:%S") - datetime.datetime(1900, 1, 1)

        new_duration = datetime.datetime.strptime(duration_str, "%H:%M:%S") - datetime.datetime(1900, 1, 1)
        percentage = (new_duration / total_duration) * 100 if total_duration.total_seconds() > 0 else 0
        return percentage

    def update_total_sheet(self, wb, current_date, duration):
        total_sheet = wb["Total"]
        date_str = current_date.strftime("%d/%m/%Y")

        for row in total_sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == date_str:
                total_duration_str = row[1].value
                total_duration = datetime.datetime.strptime(total_duration_str, "%H:%M:%S") - datetime.datetime(1900, 1, 1)
                new_duration = datetime.datetime.strptime(self._format_time(duration), "%H:%M:%S") - datetime.datetime(1900, 1, 1)
                total_duration += new_duration
                total_time_worked_str = self._format_time(total_duration.total_seconds())
                row[1].value = total_time_worked_str

    def update_percentages(self, wb, sheet_name):
        ws = wb[sheet_name]
        total_duration = datetime.timedelta()
        for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
            duration = row[5]
            if duration:
                total_duration += datetime.datetime.strptime(duration, "%H:%M:%S") - datetime.datetime(1900, 1, 1)

        for row in ws.iter_rows(min_row=2, max_col=7):
            if row[5].value:
                activity_duration = datetime.datetime.strptime(row[5].value, "%H:%M:%S") - datetime.datetime(1900, 1, 1)
                percentage = (activity_duration / total_duration) * 100 if total_duration.total_seconds() > 0 else 0
                row[6].value = f"{percentage:.2f}%"

    def _format_time(self, elapsed):
        hours, remainder = divmod(elapsed, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
